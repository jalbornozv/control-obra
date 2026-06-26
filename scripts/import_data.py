"""
Script de importación: Lee los dos archivos Excel del proyecto y carga datos
en las tablas obras, partidas de Supabase.

Requiere SUPABASE_SERVICE_KEY en .env.local (la clave service_role,
no la anon key) para poder insertar datos cuando RLS está activo.

Si no tienes la service_role key, ejecuta primero en el SQL Editor de Supabase:
  -- Permitir inserts desde el script de importación (anon key)
  CREATE POLICY "import_insert_obras" ON obras FOR INSERT TO anon WITH CHECK (true);
  CREATE POLICY "import_insert_partidas" ON partidas FOR INSERT TO anon WITH CHECK (true);
"""
import openpyxl
import requests
import json
import os
import sys
from dotenv import load_dotenv

# Load env from parent directory (project root)
load_dotenv('../.env.local')

SUPABASE_URL = os.environ['VITE_SUPABASE_URL']

# Prefer service_role key (bypasses RLS); fall back to anon key
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY') or os.environ['VITE_SUPABASE_ANON_KEY']

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=representation'
}


def supabase_post(table, data):
    r = requests.post(f'{SUPABASE_URL}/rest/v1/{table}', headers=HEADERS, json=data)
    if r.status_code == 401:
        msg = r.json().get('message', r.text)
        if 'row-level security' in r.text or '42501' in r.text:
            print('\n❌ Error RLS: La tabla tiene Row Level Security activo y la anon key no tiene permiso de INSERT.')
            print('   Solución A: Agrega SUPABASE_SERVICE_KEY=<tu service_role key> en .env.local')
            print('   Solución B: Ejecuta en el SQL Editor de Supabase:')
            print('     CREATE POLICY "import_insert_obras" ON obras FOR INSERT TO anon WITH CHECK (true);')
            print('     CREATE POLICY "import_insert_partidas" ON partidas FOR INSERT TO anon WITH CHECK (true);')
            sys.exit(1)
        raise Exception(f'401 Unauthorized: {msg}')
    r.raise_for_status()
    return r.json()


def leer_presupuesto():
    wb = openpyxl.load_workbook('../1100 Presupuesto doña carne JA.xlsx', data_only=True)
    ws = wb['Presupuesto']
    items = {}
    seccion_actual = None
    for row in ws.iter_rows(values_only=True):
        # Detectar sección (ej: "A. DEMOLICIONES Y RETIROS")
        if row[0] and isinstance(row[0], str) and len(row[0]) > 2 and row[0][1:3] == '. ':
            seccion_actual = row[0]
            continue
        num, nombre, unidad, cantidad, precio_unit, subtotal = row[0], row[1], row[2], row[3], row[4], row[5]
        if isinstance(num, (int, float)) and nombre and subtotal:
            items[str(int(num))] = {
                'seccion': seccion_actual,
                'numero': str(int(num)),
                'nombre': nombre,
                'unidad': unidad or '',
                'cantidad': float(cantidad) if cantidad else 0,
                'precio_unit': float(precio_unit) if precio_unit else 0,
                'subtotal': float(subtotal) if subtotal else 0,
            }
    return items


def leer_gantt():
    wb = openpyxl.load_workbook('../Carta_Gantt_60dias_v4.xlsx', data_only=True)
    ws = wb['Carta Gantt']
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Skip header rows (rows 0, 1, 2)
        if i < 3:
            continue
        if not row[0] or not row[3] or row[8] is None:
            continue
        cuadrilla = str(row[0])
        seccion = str(row[1]) if row[1] else ''
        numero = str(row[2]) if row[2] else ''
        nombre = str(row[3])
        unidad = str(row[4]) if row[4] else ''
        cantidad = float(row[5]) if row[5] else 0
        dia_ini = int(row[8])
        dia_fin = int(row[9]) if row[9] else dia_ini
        items.append({
            'cuadrilla': cuadrilla,
            'seccion': seccion,
            'numero': numero,
            'nombre': nombre,
            'unidad': unidad,
            'cantidad': cantidad,
            'dia_ini': dia_ini,
            'dia_fin': dia_fin,
        })
    return items


def get_presupuesto_neto():
    """Reads the subtotal neto from the Excel file."""
    wb = openpyxl.load_workbook('../1100 Presupuesto doña carne JA.xlsx', data_only=True)
    ws = wb['Presupuesto']
    for row in ws.iter_rows(values_only=True):
        if row[3] and 'Subtotal Neto' in str(row[3]):
            return float(row[5]) if row[5] else 0
    return 0


def main():
    print('Leyendo presupuesto...')
    presupuesto = leer_presupuesto()
    print(f'  {len(presupuesto)} items en presupuesto')

    print('Leyendo carta Gantt...')
    gantt = leer_gantt()
    print(f'  {len(gantt)} items en Gantt')

    presupuesto_neto = get_presupuesto_neto()
    print(f'  Presupuesto neto: ${presupuesto_neto:,.0f}')

    print('Creando obra en Supabase...')
    obra = supabase_post('obras', {
        'nombre': 'Doña Carne Manquehue 1',
        'fecha_inicio': '2026-06-26',
        'total_dias': 60,
        'presupuesto_neto': presupuesto_neto,
    })
    obra_id = obra[0]['id']
    print(f'  Obra creada: {obra_id}')

    print('Importando partidas...')
    count = 0
    for g in gantt:
        # Enrich with presupuesto data where available
        ppto = presupuesto.get(g['numero'], {})
        partida = {
            'obra_id': obra_id,
            'cuadrilla': g['cuadrilla'],
            'seccion': ppto.get('seccion') or g['seccion'],
            'numero': g['numero'],
            'nombre': g['nombre'],
            'unidad': ppto.get('unidad') or g['unidad'],
            'cantidad': ppto.get('cantidad') or g['cantidad'],
            'precio_unit': ppto.get('precio_unit', 0),
            'subtotal': ppto.get('subtotal', 0),
            'dia_ini': g['dia_ini'],
            'dia_fin': g['dia_fin'],
            'avance_pct': 0,
        }
        supabase_post('partidas', partida)
        count += 1
        print(f'  [{count}/{len(gantt)}] {g["nombre"][:60]}')

    print(f'\n✅ Importación completa: {count} partidas cargadas.')
    print(f'   Obra ID: {obra_id}')
    print(f'   Guarda este ID: lo necesitarás para la app.')


if __name__ == '__main__':
    main()
